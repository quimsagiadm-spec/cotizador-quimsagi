import streamlit as st
import pandas as pd
from supabase import create_client, Client
import io
import os
from datetime import datetime, date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from fpdf import FPDF
from PIL import Image

# --- 1. CONEXIÓN A TU BASE DE DATOS EN LA NUBE ---
URL_SUPABASE = "https://ahmjxfmfnbuhmirtfjzu.supabase.co"
CLAVE_SUPABASE = "sb_publishable_yRq3_3K7MK2d3Zxt6DmQBg_fBht0_Bv" 

@st.cache_resource
def init_connection():
    return create_client(URL_SUPABASE, CLAVE_SUPABASE)

try:
    supabase: Client = init_connection()
except Exception as e:
    st.error("Error conectando a la base de datos.")

# --- FUNCIONES PARA LEER DATOS ---
def obtener_clientes():
    try:
        return supabase.table("clientes").select("*").execute().data
    except:
        return []

def obtener_productos():
    try:
        return supabase.table("productos").select("*").execute().data
    except:
        return []

def obtener_historial():
    try:
        return supabase.table("cotizaciones").select("*").order("id", desc=True).execute().data
    except:
        return []

def obtener_siguiente_folio():
    try:
        res = supabase.table("cotizaciones").select("id").order("id", desc=True).limit(1).execute()
        if res.data:
            return res.data[0]["id"] + 1
        return 1
    except:
        return 1

# --- 2. EXCEL GENERAL ---
def generar_excel(datos_cotizacion, cliente, subtotal, iva, total, incluye_iva=True):
    wb = Workbook()
    ws = wb.active
    ws.title = "Cotización"

    header_fill = PatternFill(start_color="2B3A42", end_color="2B3A42", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    title_font = Font(size=16, bold=True, color="1F4E79")
    bold_font = Font(bold=True)
    border_thin = Border(left=Side(style='thin', color='E0E0E0'), right=Side(style='thin', color='E0E0E0'), top=Side(style='thin', color='E0E0E0'), bottom=Side(style='thin', color='E0E0E0'))

    ws['A1'] = "COTIZACIÓN QUIMSAGI"
    ws['A1'].font = title_font
    ws['A2'] = f"Cliente: {cliente}"
    ws['A2'].font = bold_font

    if datos_cotizacion:
        headers = list(datos_cotizacion[0].keys())
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for row_num, row_data in enumerate(datos_cotizacion, 5):
            for col_num, key in enumerate(headers, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = row_data[key]
                cell.border = border_thin
                if key in ["Precio Unit.", "Subtotal"]:
                    cell.number_format = '"$"#,##0.00'
                elif key == "Cant.":
                    cell.alignment = Alignment(horizontal="center")

        final_row = len(datos_cotizacion) + 5
        ws.cell(row=final_row+1, column=6).value = "Subtotal:"
        ws.cell(row=final_row+1, column=6).font = bold_font
        ws.cell(row=final_row+1, column=7).value = subtotal
        ws.cell(row=final_row+1, column=7).number_format = '"$"#,##0.00'
        
        texto_iva = "IVA (16%):" if incluye_iva else "IVA (0%):"
        ws.cell(row=final_row+2, column=6).value = texto_iva
        ws.cell(row=final_row+2, column=6).font = bold_font
        ws.cell(row=final_row+2, column=7).value = iva
        ws.cell(row=final_row+2, column=7).number_format = '"$"#,##0.00'
        
        ws.cell(row=final_row+3, column=6).value = "Total:"
        ws.cell(row=final_row+3, column=6).font = bold_font
        ws.cell(row=final_row+3, column=7).value = total
        ws.cell(row=final_row+3, column=7).number_format = '"$"#,##0.00'

    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 15

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()

# --- FILTRO LIMPIADOR ---
def limpiar_texto(texto):
    if not isinstance(texto, str):
        return str(texto)
    return texto.replace('', '').replace('\ufffd', '')

def armar_direccion(c):
    partes = [
        f"{c.get('CALLE', '')} {c.get('NO_EXTERIOR', '')}".strip(),
        f"Col. {c.get('COLONIA', '')}" if c.get('COLONIA') else "",
        c.get('MUNICIPIO', ''),
        c.get('ESTADO', ''),
        f"CP {c.get('CP', '')}" if c.get('CP') else ""
    ]
    validas = [p for p in partes if p and p.lower() != "none" and p.lower() != "nan"]
    return ", ".join(validas)

# --- 3. CLASE ESPECIAL PDF (Cotización) ---
class PDFQuimsagi(FPDF):
    def footer(self):
        self.set_y(-25)
        self.set_font("Arial", 'B', 10)
        self.set_text_color(26, 82, 118) 
        self.cell(0, 5, "Favor de confirmar la cotizacion con su vendedor", ln=True, align='C')
        self.set_font("Arial", 'B', 9)
        self.cell(0, 5, "Ventas: ventas1quimsagi@gmail.com  |  Tel: 998 459 2513", ln=True, align='C')
        self.cell(0, 5, "Administracion: direccionquimsagi@gmail.com", ln=True, align='C')

def generar_pdf(datos_cotizacion, cliente_info, subtotal, iva, total, vendedor, folio, observaciones="", forma_pago="", incluye_iva=True):
    pdf = PDFQuimsagi()
    pdf.add_page()
    
    logo_path = None
    if os.path.exists("logo.png"): logo_path = "logo.png"
    elif os.path.exists("logo.jpg"): logo_path = "logo.jpg"
    elif os.path.exists("QUIMSAGI LOGO FINAL.jpeg"): logo_path = "QUIMSAGI LOGO FINAL.jpeg"

    if logo_path:
        try:
            img = Image.open(logo_path).convert("RGBA")
            alpha = img.split()[3]
            alpha = alpha.point(lambda p: p * 0.15) 
            img.putalpha(alpha)
            wm_io = io.BytesIO()
            img.save(wm_io, format='PNG')
            wm_io.seek(0)
            pdf.image(wm_io, x=35, y=90, w=140)
        except: pass 
        try: pdf.image(logo_path, 10, 10, 42)
        except: pass
    else:
        pdf.set_font("Arial", 'B', 22)
        pdf.set_text_color(26, 82, 118)
        pdf.cell(50, 10, "Quimsagi", ln=False)

    pdf.set_font("Arial", 'B', 18)
    pdf.set_text_color(0, 0, 0)
    pdf.cell(0, 10, f"COTIZACION #{folio}", ln=True, align='R')
    
    pdf.set_font("Arial", 'I', 10)
    pdf.set_text_color(100, 100, 100)
    pdf.cell(0, 5, "PRODUCTOS DE LIMPIEZA A TU MEDIDA", ln=True, align='R')
    pdf.ln(10)
    
    pdf.set_fill_color(240, 240, 240)
    pdf.set_font("Arial", 'B', 11)
    pdf.set_text_color(0, 0, 0)
    pdf.cell(0, 8, "  Datos del Cliente", border=0, fill=True, ln=True)
    pdf.ln(2)
    
    pdf.set_font("Arial", 'B', 9)
    pdf.cell(20, 5, "Cliente:", border=0)
    pdf.set_font("Arial", '', 9)
    pdf.cell(110, 5, limpiar_texto(cliente_info.get("RAZON_SOCIAL", "")), border=0)
    
    pdf.set_font("Arial", 'B', 9)
    pdf.cell(20, 5, "Atendio:", border=0)
    pdf.set_font("Arial", '', 9)
    pdf.cell(40, 5, limpiar_texto(vendedor), border=0, ln=True)
    
    pdf.set_font("Arial", 'B', 9)
    pdf.cell(20, 5, "RFC:", border=0)
    pdf.set_font("Arial", '', 9)
    pdf.cell(110, 5, limpiar_texto(cliente_info.get("RFC", "")), border=0)
    
    pdf.set_font("Arial", 'B', 9)
    pdf.cell(20, 5, "Fecha:", border=0)
    pdf.set_font("Arial", '', 9)
    pdf.cell(40, 5, datetime.now().strftime("%d/%m/%Y"), border=0, ln=True)
    
    pdf.set_font("Arial", 'B', 9)
    pdf.cell(20, 5, "Direccion:", border=0)
    pdf.set_font("Arial", '', 9)
    pdf.multi_cell(110, 5, limpiar_texto(armar_direccion(cliente_info)), border=0)
    
    if forma_pago:
        pdf.set_font("Arial", 'B', 9)
        pdf.cell(20, 5, "Pago:", border=0)
        pdf.set_font("Arial", '', 9)
        pdf.cell(0, 5, limpiar_texto(forma_pago), border=0, ln=True)
        
    if observaciones:
        pdf.ln(2)
        pdf.set_font("Arial", 'B', 9)
        pdf.cell(25, 5, "Notas/Envio:", border=0)
        pdf.set_font("Arial", '', 9)
        pdf.multi_cell(0, 5, limpiar_texto(observaciones), border=0)
        
    pdf.ln(5)
    
    pdf.set_fill_color(26, 82, 118)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font("Arial", 'B', 9)
    
    col_w = {"Clave": 20, "Producto": 85, "Cant.": 15, "Unidad": 20, "Precio Unit.": 25, "Subtotal": 25}
    
    pdf.cell(col_w["Clave"], 8, "Clave", border=1, align='C', fill=True)
    pdf.cell(col_w["Producto"], 8, "Descripcion", border=1, align='C', fill=True)
    pdf.cell(col_w["Cant."], 8, "Cant.", border=1, align='C', fill=True)
    pdf.cell(col_w["Unidad"], 8, "Unidad", border=1, align='C', fill=True)
    pdf.cell(col_w["Precio Unit."], 8, "Precio", border=1, align='C', fill=True)
    pdf.cell(col_w["Subtotal"], 8, "Total", border=1, align='C', fill=True)
    pdf.ln()
    
    pdf.set_text_color(0, 0, 0)
    pdf.set_font("Arial", '', 9)
    
    for row in datos_cotizacion:
        pdf.cell(col_w["Clave"], 8, limpiar_texto(row["Clave"]), border=1)
        pdf.cell(col_w["Producto"], 8, limpiar_texto(row["Producto"])[:45], border=1)
        pdf.cell(col_w["Cant."], 8, str(row["Cant."]), border=1, align='C')
        pdf.cell(col_w["Unidad"], 8, limpiar_texto(row["Unidad"]), border=1, align='C')
        pdf.cell(col_w["Precio Unit."], 8, f"${row['Precio Unit.']:,.2f}", border=1, align='R')
        pdf.cell(col_w["Subtotal"], 8, f"${row['Subtotal']:,.2f}", border=1, align='R')
        pdf.ln()
        
    pdf.ln(5)
    margen_izq = col_w["Clave"] + col_w["Producto"] + col_w["Cant."] + col_w["Unidad"]
    
    pdf.set_font("Arial", '', 9)
    pdf.cell(margen_izq, 7, "", border=0)
    pdf.cell(col_w["Precio Unit."], 7, "Subtotal:", border=0, align='R')
    pdf.cell(col_w["Subtotal"], 7, f"${subtotal:,.2f}", border=0, align='R')
    pdf.ln()
    
    pdf.cell(margen_izq, 7, "", border=0)
    texto_iva = "IVA (16%):" if incluye_iva else "IVA (0%):"
    pdf.cell(col_w["Precio Unit."], 7, texto_iva, border=0, align='R')
    pdf.cell(col_w["Subtotal"], 7, f"${iva:,.2f}", border=0, align='R')
    pdf.ln()
    
    pdf.set_font("Arial", 'B', 10)
    pdf.set_fill_color(240, 240, 240)
    pdf.cell(margen_izq, 8, "", border=0)
    pdf.cell(col_w["Precio Unit."], 8, "Total:", border=1, align='R', fill=True)
    pdf.cell(col_w["Subtotal"], 8, f"${total:,.2f}", border=1, align='R', fill=True)
    pdf.ln(15)
    
    pdf.set_font("Arial", 'B', 10)
    pdf.set_text_color(26, 82, 118) 
    pdf.cell(0, 6, "DATOS PARA TRANSFERENCIA BANCARIA", ln=True)
    
    pdf.set_font("Arial", 'B', 9) 
    pdf.cell(15, 5, "Banco:", border=0)
    pdf.cell(80, 5, "Mercado Pago", border=0, ln=True)
    
    pdf.cell(15, 5, "Titular:", border=0)
    pdf.cell(80, 5, "Quimsagi", border=0, ln=True)
    
    pdf.cell(15, 5, "CLABE:", border=0)
    pdf.cell(80, 5, "722969013491074912", border=0, ln=True)

    pdf_bytes = pdf.output(dest='S')
    return bytes(pdf_bytes) if not isinstance(pdf_bytes, str) else pdf_bytes.encode('latin-1')

# --- 4. GENERADOR PDF ESTADO DE CUENTA CLIENTE ---
def generar_estado_cuenta_pdf(nombre_cliente, cliente_info, lista_facturas_cliente):
    pdf = FPDF()
    pdf.add_page()
    
    logo_path = None
    if os.path.exists("logo.png"): logo_path = "logo.png"
    elif os.path.exists("logo.jpg"): logo_path = "logo.jpg"
    elif os.path.exists("QUIMSAGI LOGO FINAL.jpeg"): logo_path = "QUIMSAGI LOGO FINAL.jpeg"

    if logo_path:
        try: pdf.image(logo_path, 10, 10, 35)
        except: pass

    pdf.set_font("Arial", 'B', 16)
    pdf.set_text_color(26, 82, 118)
    pdf.cell(0, 10, "ESTADO DE CUENTA", ln=True, align='R')
    pdf.set_font("Arial", 'I', 9)
    pdf.set_text_color(100, 100, 100)
    pdf.cell(0, 4, "QUIMSAGI - PRODUCTOS DE LIMPIEZA", ln=True, align='R')
    pdf.ln(10)
    
    pdf.set_fill_color(240, 240, 240)
    pdf.set_font("Arial", 'B', 10)
    pdf.set_text_color(0, 0, 0)
    pdf.cell(0, 7, "  Informacion del Cliente", border=0, fill=True, ln=True)
    pdf.ln(2)
    
    pdf.set_font("Arial", 'B', 9)
    pdf.cell(20, 5, "Cliente:", border=0)
    pdf.set_font("Arial", '', 9)
    pdf.cell(100, 5, limpiar_texto(nombre_cliente), border=0, ln=True)
    
    pdf.set_font("Arial", 'B', 9)
    pdf.cell(20, 5, "RFC:", border=0)
    pdf.set_font("Arial", '', 9)
    pdf.cell(100, 5, limpiar_texto(cliente_info.get("RFC", "N/D")), border=0, ln=True)
    
    pdf.set_font("Arial", 'B', 9)
    pdf.cell(20, 5, "Direccion:", border=0)
    pdf.set_font("Arial", '', 9)
    pdf.multi_cell(0, 5, limpiar_texto(armar_direccion(cliente_info)), border=0)
    pdf.ln(10)
    
    pdf.set_fill_color(26, 82, 118)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font("Arial", 'B', 9)
    
    pdf.cell(15, 8, "Folio", border=1, align='C', fill=True)
    pdf.cell(25, 8, "Fecha", border=1, align='C', fill=True)
    pdf.cell(35, 8, "Folio Fiscal", border=1, align='C', fill=True)
    pdf.cell(30, 8, "Operativo", border=1, align='C', fill=True)
    pdf.cell(30, 8, "Financiero", border=1, align='C', fill=True)
    pdf.cell(20, 8, "Atraso", border=1, align='C', fill=True)
    pdf.cell(25, 8, "Total", border=1, align='C', fill=True)
    pdf.ln()
    
    pdf.set_text_color(0, 0, 0)
    pdf.set_font("Arial", '', 8)
    
    suma_total_deuda = 0
    for fac in lista_facturas_cliente:
        f_creacion = pd.to_datetime(fac.get("fecha"))
        dias = (datetime.now() - f_creacion.tz_localize(None) if f_creacion.tzinfo else datetime.now() - f_creacion).days
        if fac.get("estatus_financiero") == "Pagada":
            dias_str = "Pagado"
        else:
            dias_str = f"{dias} dias"
            suma_total_deuda += float(fac.get("total", 0))
            
        pdf.cell(15, 7, str(fac.get("id")), border=1, align='C')
        pdf.cell(25, 7, f_creacion.strftime('%d/%m/%Y'), border=1, align='C')
        pdf.cell(35, 7, limpiar_texto(fac.get("folio_fiscal") or "S/F"), border=1, align='C')
        pdf.cell(30, 7, limpiar_texto(fac.get("estatus_operativo", "Pendiente")), border=1, align='C')
        pdf.cell(30, 7, limpiar_texto(fac.get("estatus_financiero", "Pendiente")), border=1, align='C')
        pdf.cell(20, 7, dias_str, border=1, align='C')
        pdf.cell(25, 7, f"${float(fac.get('total', 0)):,.2f}", border=1, align='R')
        pdf.ln()
        
    pdf.ln(5)
    pdf.set_font("Arial", 'B', 10)
    pdf.cell(135, 8, "SALDO PENDIENTE TOTAL:", border=0, align='R')
    pdf.cell(25, 8, f"${suma_total_deuda:,.2f}", border=1, align='R', fill=True)

    pdf_bytes = pdf.output(dest='S')
    return bytes(pdf_bytes) if not isinstance(pdf_bytes, str) else pdf_bytes.encode('latin-1')

# --- PÁGINA PRINCIPAL ---
st.set_page_config(page_title="ERP QUIMSAGI", page_icon="📋", layout="wide")

# --- VARIABLES DE SESIÓN ---
if 'cotizacion_actual' not in st.session_state: st.session_state.cotizacion_actual = []
if 'autenticado' not in st.session_state: st.session_state.autenticado = False
if 'folio_en_edicion' not in st.session_state: st.session_state.folio_en_edicion = None
if 'cliente_en_edicion' not in st.session_state: st.session_state.cliente_en_edicion = None
if 'vendedor_en_edicion' not in st.session_state: st.session_state.vendedor_en_edicion = None
if 'obs_en_edicion' not in st.session_state: st.session_state.obs_en_edicion = ""
if 'fp_cot_en_edicion' not in st.session_state: st.session_state.fp_cot_en_edicion = "Transferencia"
if 'incluye_iva_en_edicion' not in st.session_state: st.session_state.incluye_iva_en_edicion = True

if not st.session_state.autenticado:
    st.title("🔒 Acceso al Sistema")
    usuario = st.text_input("Usuario")
    contrasena = st.text_input("Contraseña", type="password")
    
    if st.button("Entrar"):
        if usuario == "admin" and contrasena == "12345":
            st.session_state.autenticado = True
            st.rerun()
        else:
            st.error("Usuario o contraseña incorrectos")
else:
    st.sidebar.title("Menú Principal")
    menu = st.sidebar.radio("Ir a:", ["📝 Cotizador", "📂 Historial y Cobranza", "📊 Métricas", "⚙️ Panel de Administrador"])
    
    if st.sidebar.button("Cerrar Sesión"):
        st.session_state.autenticado = False
        st.rerun()

    # ==========================================
    # 1. MÓDULO COTIZADOR
    # ==========================================
    if menu == "📝 Cotizador":
        st.title("📝 Módulo de Ventas - ERP QUIMSAGI")
        clientes = obtener_clientes()
        productos = obtener_productos()
        
        if len(clientes) == 0 or len(productos) == 0:
            st.warning("⚠️ Faltan clientes o productos en la base de datos.")
        else:
            if st.session_state.folio_en_edicion:
                st.warning(f"⚠️ ESTÁS EDITANDO EL FOLIO #{st.session_state.folio_en_edicion} DESDE EL HISTORIAL")
                if st.button("❌ Cancelar edición y hacer cotización nueva", use_container_width=True):
                    st.session_state.folio_en_edicion = None
                    st.session_state.cliente_en_edicion = None
                    st.session_state.vendedor_en_edicion = None
                    st.session_state.obs_en_edicion = ""
                    st.session_state.fp_cot_en_edicion = "Transferencia"
                    st.session_state.incluye_iva_en_edicion = True
                    st.session_state.cotizacion_actual = []
                    st.rerun()
            
            folio_actual = st.session_state.folio_en_edicion if st.session_state.folio_en_edicion else obtener_siguiente_folio()
            
            st.subheader("1. Datos Generales")
            col_vendedor, col_folio = st.columns([3, 1])
            
            lista_vendedores = ["Gilber Carbajal", "Omar Santiago", "Lizedy Facundo", "Grisy Ojeda"]
            idx_vendedor = lista_vendedores.index(st.session_state.vendedor_en_edicion) if st.session_state.vendedor_en_edicion in lista_vendedores else 0
            
            lista_razones_sociales = [c.get("RAZON_SOCIAL", "Sin Razón Social") for c in clientes if c.get("RAZON_SOCIAL")]
            idx_cliente = lista_razones_sociales.index(st.session_state.cliente_en_edicion) if st.session_state.cliente_en_edicion in lista_razones_sociales else 0

            with col_vendedor:
                vendedor_seleccionado = st.selectbox("Atendido por:", lista_vendedores, index=idx_vendedor)
            with col_folio:
                st.info(f"Folio: **#{folio_actual}**")

            col_c1, col_c2 = st.columns([2, 1])
            with col_c1:
                cliente_seleccionado = st.selectbox("Selecciona un cliente:", lista_razones_sociales, index=idx_cliente)
            with col_c2:
                opciones_fp = ["Transferencia", "Efectivo", "Tarjeta", "Cheque", "Crédito"]
                idx_fp = opciones_fp.index(st.session_state.fp_cot_en_edicion) if st.session_state.fp_cot_en_edicion in opciones_fp else 0
                fp_cotizacion_seleccionada = st.selectbox("Forma de pago esperada:", opciones_fp, index=idx_fp)
                
                # --- NUEVO: CHECKBOX DE IVA ---
                aplicar_iva = st.checkbox("☑️ Incluir IVA (16%)", value=st.session_state.get('incluye_iva_en_edicion', True))

            observaciones_texto = st.text_area("Observaciones o Sucursal de entrega (Se imprimirá en el PDF):", value=st.session_state.obs_en_edicion, height=68)
            
            st.subheader("2. Agregar Productos y Precios")
            mapa_productos = {f"{p.get('CLAVE') or 'S/C'} - {p.get('PRODUCTO', 'Sin Nombre')}": p for p in productos if p.get("PRODUCTO")}
            lista_opciones_productos = list(mapa_productos.keys())
            
            col1, col2, col3 = st.columns([2, 1, 1])
            with col1: 
                producto_seleccionado_display = st.selectbox("Selecciona un producto:", lista_opciones_productos)
            
            prod_info = mapa_productos.get(producto_seleccionado_display)
            producto_seleccionado = prod_info.get("PRODUCTO") if prod_info else ""
            precio_catalogo = float(prod_info.get("PRECIO", 0) or 0) if prod_info else 0.0
            
            with col2: 
                cantidad = st.number_input("Cantidad", min_value=1, value=1)
            with col3:
                precio_personalizado = st.number_input("Precio Unitario ($)", min_value=0.0, value=precio_catalogo, format="%.2f")
                
            if st.button("Agregar a la cotización"):
                if prod_info:
                    descuento_pct = float(prod_info.get("DESCUENTO", 0) or 0)
                    if precio_personalizado != precio_catalogo:
                        precio_final_unitario = precio_personalizado
                    else:
                        descuento_dinero = precio_catalogo * (descuento_pct / 100)
                        precio_final_unitario = precio_catalogo - descuento_dinero
                        
                    subtotal_linea = precio_final_unitario * cantidad
                    st.session_state.cotizacion_actual.append({
                        "Clave": prod_info.get("Clave", prod_info.get("CLAVE", "")),
                        "Producto": producto_seleccionado,
                        "Cant.": cantidad,
                        "Unidad": prod_info.get("UNIDAD", "Pza"),
                        "Precio Unit.": precio_final_unitario, 
                        "Desc. %": f"{descuento_pct}%",
                        "Subtotal": subtotal_linea
                    })
                    st.success(f"Se agregó {cantidad}x {producto_seleccionado} a ${precio_final_unitario:,.2f}")

            if len(st.session_state.cotizacion_actual) > 0:
                st.markdown("---")
                st.subheader("🛒 Resumen de la Cotización")
                df = pd.DataFrame(st.session_state.cotizacion_actual)
                df_vista = df.copy()
                df_vista["Precio Unit."] = df_vista["Precio Unit."].apply(lambda x: f"${x:,.2f}")
                df_vista["Subtotal"] = df_vista["Subtotal"].apply(lambda x: f"${x:,.2f}")
                st.dataframe(df_vista, use_container_width=True, hide_index=True)
                
                st.markdown("#### 🛠️ Modificar Carrito")
                tab_edit, tab_del = st.tabs(["✏️ Editar Producto", "❌ Eliminar Producto"])
                opciones_carrito = [f"Línea {i+1}: {prod['Producto']} (Cant: {prod['Cant.']})" for i, prod in enumerate(st.session_state.cotizacion_actual)]
                
                with tab_edit:
                    col_sel, col_cant, col_precio, col_btn = st.columns([2, 1, 1, 1])
                    with col_sel: item_a_editar = st.selectbox("Selecciona para editar:", opciones_carrito, key="sel_edit")
                    if item_a_editar:
                        idx_edit = opciones_carrito.index(item_a_editar)
                        prod_edit = st.session_state.cotizacion_actual[idx_edit]
                        with col_cant: nueva_cant = st.number_input("Cantidad", min_value=1, value=int(prod_edit['Cant.']), key="cant_edit")
                        with col_precio: nuevo_precio = st.number_input("Precio ($)", min_value=0.0, value=float(prod_edit['Precio Unit.']), format="%.2f", key="precio_edit")
                        with col_btn:
                            st.markdown("<br>", unsafe_allow_html=True)
                            if st.button("✏️ Actualizar", use_container_width=True):
                                st.session_state.cotizacion_actual[idx_edit]['Cant.'] = nueva_cant
                                st.session_state.cotizacion_actual[idx_edit]['Precio Unit.'] = nuevo_precio
                                st.session_state.cotizacion_actual[idx_edit]['Subtotal'] = nueva_cant * nuevo_precio
                                st.success("¡Línea actualizada exitosamente!")
                                st.rerun()

                with tab_del:
                    col_del1, col_del2 = st.columns([3, 1])
                    with col_del1: item_a_borrar = st.selectbox("Selecciona para eliminar:", opciones_carrito, key="sel_del")
                    with col_del2:
                        st.markdown("<br>", unsafe_allow_html=True)
                        if st.button("❌ Eliminar seleccionado", use_container_width=True):
                            idx_del = opciones_carrito.index(item_a_borrar)
                            st.session_state.cotizacion_actual.pop(idx_del)
                            st.success("Producto eliminado del carrito.")
                            st.rerun()
                
                # --- NUEVO: CÁLCULO DINÁMICO DEL IVA ---
                suma_subtotal = df["Subtotal"].sum() if len(st.session_state.cotizacion_actual) > 0 else 0
                iva = suma_subtotal * 0.16 if aplicar_iva else 0.0
                total_final = suma_subtotal + iva
                
                if len(st.session_state.cotizacion_actual) > 0:
                    st.markdown("---")
                    col_blank, col_totales = st.columns([3, 1])
                    with col_totales:
                        st.write(f"**Subtotal:** ${suma_subtotal:,.2f}")
                        texto_iva_ui = "**IVA (16%):**" if aplicar_iva else "**IVA (0%):**"
                        st.write(f"{texto_iva_ui} ${iva:,.2f}")
                        st.write(f"### **Total:** ${total_final:,.2f}")
                    
                    st.markdown("---")
                    col_btn1, col_btn2, col_btn3, col_btn4 = st.columns(4)
                    cliente_completo = next((c for c in clientes if c.get("RAZON_SOCIAL") == cliente_seleccionado), {})
                    
                    with col_btn1:
                        texto_boton_guardar = "💾 Actualizar Folio" if st.session_state.folio_en_edicion else "💾 Guardar en Historial"
                        if st.button(texto_boton_guardar, use_container_width=True):
                            cotizacion_data = {
                                "cliente": cliente_seleccionado, 
                                "vendedor": vendedor_seleccionado,
                                "total": total_final,
                                "detalles": st.session_state.cotizacion_actual,
                                "estatus_operativo": "Pendiente de autorización",
                                "observaciones": observaciones_texto,
                                "forma_pago_cotizacion": fp_cotizacion_seleccionada,
                                "incluye_iva": aplicar_iva
                            }
                            
                            try:
                                if st.session_state.folio_en_edicion:
                                    supabase.table("cotizaciones").update(cotizacion_data).eq("id", st.session_state.folio_en_edicion).execute()
                                    st.success(f"¡Folio #{st.session_state.folio_en_edicion} actualizado exitosamente!")
                                else:
                                    cotizacion_data["estatus_financiero"] = "Pendiente de cobro"
                                    supabase.table("cotizaciones").insert(cotizacion_data).execute()
                                    st.success("¡Guardada exitosamente!")
                                    
                                st.session_state.cotizacion_actual = []
                                st.session_state.folio_en_edicion = None
                                st.session_state.cliente_en_edicion = None
                                st.session_state.vendedor_en_edicion = None
                                st.session_state.obs_en_edicion = ""
                                st.session_state.fp_cot_en_edicion = "Transferencia"
                                st.session_state.incluye_iva_en_edicion = True
                                st.rerun()
                            except Exception as e: 
                                st.error(f"Error al guardar: {e}")
                                
                    with col_btn2:
                        st.download_button("📥 Descargar Excel", data=generar_excel(st.session_state.cotizacion_actual, cliente_seleccionado, suma_subtotal, iva, total_final, aplicar_iva), file_name=f"Cotizacion_Folio{folio_actual}_{cliente_seleccionado}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
                    with col_btn3:
                        st.download_button("📄 Descargar PDF", data=generar_pdf(st.session_state.cotizacion_actual, cliente_completo, suma_subtotal, iva, total_final, vendedor_seleccionado, folio_actual, observaciones_texto, fp_cotizacion_seleccionada, aplicar_iva), file_name=f"Cotizacion_Folio{folio_actual}_{cliente_seleccionado}.pdf", mime="application/pdf", use_container_width=True)
                    with col_btn4:
                        if st.button("🗑️ Limpiar Todo", use_container_width=True):
                            st.session_state.cotizacion_actual = []
                            st.rerun()

    # ==========================================
    # 2. HISTORIAL Y COBRANZA
    # ==========================================
    elif menu == "📂 Historial y Cobranza":
        st.title("📂 Historial y Cuentas por Cobrar (CxC)")
        historial = obtener_historial()
        if len(historial) == 0:
            st.info("Aún no hay cotizaciones guardadas.")
        else:
            df_hist = pd.DataFrame(historial)
            for col, val in [("vendedor", "S/D"), ("estatus_operativo", "Pendiente de autorización"), ("estatus_financiero", "Pendiente de cobro"), ("folio_fiscal", "-"), ("forma_pago", "-"), ("observaciones", ""), ("forma_pago_cotizacion", "")]:
                if col not in df_hist.columns: df_hist[col] = val
                else: df_hist[col] = df_hist[col].fillna(val)

            dias_atraso_lista = []
            for idx, row in df_hist.iterrows():
                f_creacion = pd.to_datetime(row.get("fecha"))
                dias = (datetime.now() - f_creacion.tz_localize(None) if f_creacion.tzinfo else datetime.now() - f_creacion).days
                if row.get("estatus_operativo") == "No Autorizada": dias_atraso_lista.append("Cancelada")
                elif row.get("estatus_financiero") == "Pagada": dias_atraso_lista.append("Pagado")
                else: dias_atraso_lista.append(f"{dias} días")
            df_hist["Atraso"] = dias_atraso_lista

            df_vista = df_hist[["id", "cliente", "vendedor", "estatus_operativo", "estatus_financiero", "Atraso", "total", "fecha"]].copy()
            df_vista["total"] = df_vista["total"].apply(lambda x: f"${float(x):,.2f}")
            df_vista["fecha"] = pd.to_datetime(df_vista["fecha"]).dt.strftime('%d/%m/%Y %H:%M')
            st.dataframe(df_vista, column_config={"id": "Folio", "cliente": "Cliente", "vendedor": "Atendió", "estatus_operativo": "Operativo", "estatus_financiero": "Financiero", "Atraso": "Antigüedad", "total": "Total", "fecha": "Fecha"}, use_container_width=True, hide_index=True)

            st.markdown("---")
            st.subheader("⚡ Administrar Estatus y Cobranza de un Folio")
            opciones_folios = [f"Folio #{r['id']} - {r['cliente']}" for r in historial]
            
            folio_sel_str = st.selectbox("Selecciona el folio a modificar:", opciones_folios, key="memoria_folio_admin")
            
            id_sel = int(folio_sel_str.split("#")[1].split(" -")[0])
            reg_sel = next((r for r in historial if r["id"] == id_sel), {})

            col_c1, col_c2 = st.columns(2)
            with col_c1:
                st.markdown("##### 📌 Carril Operativo")
                opciones_op = ["Pendiente de autorización", "Autorizado", "Facturado", "No Autorizada"]
                op_actual = reg_sel.get("estatus_operativo")
                idx_op = opciones_op.index(op_actual) if op_actual in opciones_op else 0
                nuevo_op = st.selectbox("Estatus Operativo:", opciones_op, index=idx_op)
                nuevo_folio_fiscal = reg_sel.get("folio_fiscal") or ""
                if nuevo_op == "Facturado":
                    nuevo_folio_fiscal = st.text_input("Folio Fiscal de la Factura:", value=nuevo_folio_fiscal)

            with col_c2:
                st.markdown("##### 💰 Carril Financiero (CxC)")
                opciones_fin = ["Pendiente de cobro", "Pagada"]
                fin_actual = reg_sel.get("estatus_financiero", "Pendiente de cobro")
                idx_fin = opciones_fin.index(fin_actual) if fin_actual in opciones_fin else 0
                nuevo_fin = st.selectbox("Estatus Financiero:", opciones_fin, index=idx_fin)
                nueva_forma_pago = "Transferencia"
                nueva_fecha_pago = str(date.today())
                
                if nuevo_fin == "Pagada":
                    opciones_pago = ["Transferencia", "Efectivo", "Tarjeta", "Cheque"]
                    fp_actual = reg_sel.get("forma_pago")
                    idx_fp = opciones_pago.index(fp_actual) if fp_actual in opciones_pago else 0
                    nueva_forma_pago = st.selectbox("Forma de Pago del depósito:", opciones_pago, index=idx_fp)
                    fe_actual = reg_sel.get("fecha_pago")
                    try: dt_val = datetime.strptime(str(fe_actual), "%Y-%m-%d").date() if fe_actual else date.today()
                    except: dt_val = date.today()
                    nueva_fecha_pago = str(st.date_input("Fecha en que se pagó:", value=dt_val))

            if st.button("💾 Guardar Cambios de Estatus", use_container_width=True):
                datos_actualizados = {
                    "estatus_operativo": nuevo_op,
                    "folio_fiscal": nuevo_folio_fiscal if nuevo_op == "Facturado" else None,
                    "estatus_financiero": nuevo_fin,
                    "forma_pago": nueva_forma_pago if nuevo_fin == "Pagada" else None,
                    "fecha_pago": nueva_fecha_pago if nuevo_fin == "Pagada" else None
                }
                try:
                    supabase.table("cotizaciones").update(datos_actualizados).eq("id", id_sel).execute()
                    st.success("¡Estatus y cobranza actualizados correctamente!")
                    st.rerun()
                except Exception as e:
                    st.error(f"Error al actualizar: {e}")

            st.markdown("---")
            st.markdown("##### 🔄 Modificar Productos del Folio (Viaje en el tiempo)")
            if st.button("✏️ Cargar este Folio al Cotizador para editar productos", use_container_width=True):
                st.session_state.folio_en_edicion = reg_sel["id"]
                st.session_state.cliente_en_edicion = reg_sel["cliente"]
                st.session_state.vendedor_en_edicion = reg_sel.get("vendedor", "Gilber Carbajal")
                st.session_state.obs_en_edicion = reg_sel.get("observaciones", "")
                st.session_state.fp_cot_en_edicion = reg_sel.get("forma_pago_cotizacion", "Transferencia")
                
                # --- NUEVO: RECUPERAR EL ESTATUS DE IVA ---
                iva_bd = reg_sel.get("incluye_iva")
                st.session_state.incluye_iva_en_edicion = True if iva_bd is None else iva_bd
                
                st.session_state.cotizacion_actual = reg_sel.get("detalles", []) if reg_sel.get("detalles") else []
                st.success("✅ ¡Folio cargado con éxito! Haz clic en '📝 Cotizador' en el menú de la izquierda para ver y editar los productos.")
            
            st.markdown("---")
            col_desc1, col_desc2 = st.columns(2)
            
            with col_desc1:
                st.subheader("📄 Recuperar PDF de Cotización")
                folios_detalles = [f"Folio #{r['id']} - {r['cliente']}" for r in historial if r.get("detalles")]
                if folios_detalles:
                    f_dl = st.selectbox("Elige la cotización:", folios_detalles)
                    id_dl = int(f_dl.split("#")[1].split(" -")[0])
                    reg_dl = next((r for r in historial if r["id"] == id_dl), None)
                    if reg_dl and reg_dl.get("detalles"):
                        cli_comp = next((c for c in obtener_clientes() if c.get("RAZON_SOCIAL") == reg_dl["cliente"]), {"RAZON_SOCIAL": reg_dl["cliente"]})
                        dt_cot = reg_dl["detalles"]
                        s_sub = sum([float(i["Subtotal"]) for i in dt_cot])
                        
                        # --- NUEVO: RECALCULAR IVA EN EL HISTORIAL PARA PDF ---
                        iva_recuperado = reg_dl.get("incluye_iva")
                        iva_recuperado = True if iva_recuperado is None else iva_recuperado
                        iv = s_sub * 0.16 if iva_recuperado else 0.0
                        
                        obs_rec = reg_dl.get("observaciones", "")
                        fp_rec = reg_dl.get("forma_pago_cotizacion", "")
                        
                        pdf_rec = generar_pdf(dt_cot, cli_comp, s_sub, iv, float(reg_dl["total"]), reg_dl.get("vendedor", "Gilber Carbajal"), id_dl, obs_rec, fp_rec, iva_recuperado)
                        st.download_button("📥 Descargar PDF", data=pdf_rec, file_name=f"Cotizacion_Folio{id_dl}_{reg_dl['cliente']}.pdf", mime="application/pdf", use_container_width=True)

            with col_desc2:
                st.subheader("📑 Estado de Cuenta por Cliente")
                lista_nombres_clientes = list(set([r["cliente"] for r in historial]))
                if lista_nombres_clientes:
                    cli_edo_cta = st.selectbox("Selecciona cliente para estado de cuenta:", lista_nombres_clientes)
                    facturas_cliente = [r for r in historial if r["cliente"] == cli_edo_cta and r.get("estatus_operativo") != "No Autorizada"]
                    if len(facturas_cliente) > 0:
                        cli_info_completo = next((c for c in obtener_clientes() if c.get("RAZON_SOCIAL") == cli_edo_cta), {"RAZON_SOCIAL": cli_edo_cta})
                        pdf_edo_cta = generar_estado_cuenta_pdf(cli_edo_cta, cli_info_completo, facturas_cliente)
                        st.download_button("📥 Descargar Estado de Cuenta (PDF)", data=pdf_edo_cta, file_name=f"EstadoDeCuenta_{cli_edo_cta}.pdf", mime="application/pdf", use_container_width=True)
                    else:
                        st.info("Este cliente no tiene facturas pendientes o activas.")

    # ==========================================
    # 3. MÉTRICAS
    # ==========================================
    elif menu == "📊 Métricas":
        st.title("📊 Panel de Inteligencia Financiera")
        historial = obtener_historial()
        if len(historial) > 0:
            df = pd.DataFrame(historial)
            df['total'] = df['total'].astype(float)
            for col, val in [("estatus_operativo", "Pendiente de autorización"), ("estatus_financiero", "Pendiente de cobro"), ("vendedor", "S/D")]:
                if col not in df.columns: df[col] = val
                else: df[col] = df[col].fillna(val)

            df_validas = df[df['estatus_operativo'] != 'No Autorizada']
            col1, col2, col3, col4 = st.columns(4)
            t_historico = df_validas['total'].sum()
            t_cobrado = df_validas[df_validas['estatus_financiero'] == 'Pagada']['total'].sum()
            t_por_cobrar = df_validas[df_validas['estatus_financiero'] == 'Pendiente de cobro']['total'].sum()
            t_autorizado = df_validas[df_validas['estatus_operativo'].isin(['Autorizado', 'Facturado'])]['total'].sum()
            
            col1.metric("Volumen Venta Real", f"${t_historico:,.2f}")
            col2.metric("Cobrado en Banco", f"${t_cobrado:,.2f}")
            col3.metric("Por Cobrar (CxC)", f"${t_por_cobrar:,.2f}", delta_color="inverse")
            col4.metric("Operativo Autorizado", f"${t_autorizado:,.2f}")
            st.markdown("---")
            col_g1, col_g2 = st.columns(2)
            with col_g1:
                st.subheader("Ventas por Vendedor (Reales)")
                st.bar_chart(df_validas.groupby("vendedor")["total"].sum())
            with col_g2:
                st.subheader("Estado Financiero (Cobranza)")
                st.bar_chart(df_validas.groupby("estatus_financiero")["total"].sum())
                
            st.markdown("---")
            st.subheader("Top Clientes con mayor deuda (Pendiente de cobro)")
            df_deuda = df_validas[df_validas['estatus_financiero'] == 'Pendiente de cobro']
            if len(df_deuda) > 0:
                top_deuda = df_deuda.groupby("cliente")["total"].sum().sort_values(ascending=False).head(5).reset_index()
                top_deuda["total"] = top_deuda["total"].apply(lambda x: f"${x:,.2f}")
                st.dataframe(top_deuda, column_config={"cliente": "Cliente", "total": "Deuda Pendiente"}, use_container_width=True, hide_index=True)
            else: st.success("¡Excelente! No hay saldos pendientes de cobro en este momento.")
        else: st.info("Aún no hay suficientes datos para mostrar métricas.")

    # ==========================================
    # 4. ADMINISTRACIÓN
    # ==========================================
    elif menu == "⚙️ Panel de Administrador":
        st.title("⚙️ Configuración del ERP")
        tab1, tab2 = st.tabs(["📦 Catálogo de Productos", "🏢 Catálogo de Clientes"])
        
        # --- PRODUCTOS ---
        with tab1:
            st.subheader("Gestión de Productos")
            sub_p1, sub_p2 = st.tabs(["Crear Nuevo", "Editar Existente"])
            with sub_p1:
                col1, col2 = st.columns(2)
                with col1:
                    clave = st.text_input("CLAVE")
                    producto = st.text_input("PRODUCTO")
                    unidad = st.selectbox("UNIDAD", ["Pza", "Kg", "Litro", "Galón", "Servicio"])
                with col2:
                    precio = st.number_input("PRECIO", min_value=0.0, format="%.2f")
                    descuento = st.number_input("DESCUENTO (%)", min_value=0.0, format="%.2f")
                if st.button("Guardar Producto Nuevo"):
                    nuevo_producto = {"CLAVE": clave, "PRODUCTO": producto, "PRECIO": precio, "DESCUENTO": descuento, "UNIDAD": unidad}
                    try:
                        supabase.table("productos").insert(nuevo_producto).execute()
                        st.success("¡Producto guardado exitosamente!")
                    except: st.error("Error al guardar.")
            
            with sub_p2:
                lista_prod_admin = obtener_productos()
                if lista_prod_admin:
                    mapa_p_admin = {f"{p.get('id', p.get('CLAVE', ''))} - {p.get('PRODUCTO')}": p for p in lista_prod_admin if p.get("PRODUCTO")}
                    sel_p_str = st.selectbox("Seleccionar producto para editar:", list(mapa_p_admin.keys()))
                    
                    if sel_p_str:
                        prod_data = mapa_p_admin[sel_p_str]
                        id_prod = prod_data.get("id")
                        
                        col_e1, col_e2 = st.columns(2)
                        with col_e1:
                            e_clave = st.text_input("CLAVE", value=prod_data.get("CLAVE", ""))
                            e_producto = st.text_input("PRODUCTO", value=prod_data.get("PRODUCTO", ""))
                            u_actual = prod_data.get("UNIDAD", "Pza")
                            e_unidad = st.selectbox("UNIDAD", ["Pza", "Kg", "Litro", "Galón", "Servicio"], index=["Pza", "Kg", "Litro", "Galón", "Servicio"].index(u_actual) if u_actual in ["Pza", "Kg", "Litro", "Galón", "Servicio"] else 0, key="u_edit")
                        with col_e2:
                            e_precio = st.number_input("PRECIO", min_value=0.0, value=float(prod_data.get("PRECIO", 0) or 0), format="%.2f", key="p_edit")
                            e_descuento = st.number_input("DESCUENTO (%)", min_value=0.0, value=float(prod_data.get("DESCUENTO", 0) or 0), format="%.2f", key="d_edit")
                        
                        if st.button("Actualizar Producto"):
                            datos_act_p = {"CLAVE": e_clave, "PRODUCTO": e_producto, "PRECIO": e_precio, "DESCUENTO": e_descuento, "UNIDAD": e_unidad}
                            try:
                                if id_prod:
                                    supabase.table("productos").update(datos_act_p).eq("id", id_prod).execute()
                                else:
                                    supabase.table("productos").update(datos_act_p).eq("CLAVE", prod_data.get("CLAVE")).execute()
                                st.success("¡Producto actualizado exitosamente!")
                                st.rerun()
                            except Exception as e: st.error(f"Error al actualizar: {e}")
                else: st.info("No hay productos registrados.")

        # --- CLIENTES ---
        with tab2:
            st.subheader("Gestión de Clientes")
            sub_c1, sub_c2 = st.tabs(["Crear Nuevo", "Editar Existente"])
            
            with sub_c1:
                rfc = st.text_input("RFC")
                razon_social = st.text_input("RAZÓN SOCIAL")
                forma_pago = st.selectbox("FORMA DE PAGO HABITUAL", ["Transferencia", "Efectivo", "Tarjeta", "Crédito"])
                
                st.markdown("**Dirección**")
                col_d1, col_d2 = st.columns(2)
                with col_d1:
                    calle = st.text_input("Calle")
                    no_ext = st.text_input("No. Exterior")
                    colonia = st.text_input("Colonia")
                with col_d2:
                    municipio = st.text_input("Municipio / Ciudad")
                    estado = st.text_input("Estado")
                    cp = st.text_input("Código Postal")
                    
                if st.button("Guardar Cliente Nuevo"):
                    nuevo_cliente = {"RFC": rfc, "RAZON_SOCIAL": razon_social, "FORMA_PAGO": forma_pago, "CALLE": calle, "NO_EXTERIOR": no_ext, "COLONIA": colonia, "MUNICIPIO": municipio, "ESTADO": estado, "CP": cp}
                    try:
                        supabase.table("clientes").insert(nuevo_cliente).execute()
                        st.success("¡Cliente guardado exitosamente!")
                    except: st.error("Error al guardar.")
            
            with sub_c2:
                lista_cli_admin = obtener_clientes()
                if lista_cli_admin:
                    mapa_c_admin = {f"{c.get('RAZON_SOCIAL', 'Sin Nombre')} (RFC: {c.get('RFC', 'S/D')})": c for c in lista_cli_admin if c.get("RAZON_SOCIAL")}
                    sel_c_str = st.selectbox("Seleccionar cliente para editar:", list(mapa_c_admin.keys()))
                    
                    if sel_c_str:
                        cli_data = mapa_c_admin[sel_c_str]
                        id_cli = cli_data.get("id")
                        
                        e_rfc = st.text_input("RFC", value=cli_data.get("RFC", ""), key="c_rfc")
                        e_razon_social = st.text_input("RAZÓN SOCIAL", value=cli_data.get("RAZON_SOCIAL", ""), key="c_rs")
                        fp_act = cli_data.get("FORMA_PAGO", "Transferencia")
                        e_forma_pago = st.selectbox("FORMA DE PAGO HABITUAL", ["Transferencia", "Efectivo", "Tarjeta", "Crédito"], index=["Transferencia", "Efectivo", "Tarjeta", "Crédito"].index(fp_act) if fp_act in ["Transferencia", "Efectivo", "Tarjeta", "Crédito"] else 0, key="c_fp")
                        
                        st.markdown("**Dirección**")
                        col_ed1, col_ed2 = st.columns(2)
                        with col_ed1:
                            e_calle = st.text_input("Calle", value=cli_data.get("CALLE", ""), key="c_calle")
                            e_no_ext = st.text_input("No. Exterior", value=cli_data.get("NO_EXTERIOR", ""), key="c_next")
                            e_colonia = st.text_input("Colonia", value=cli_data.get("COLONIA", ""), key="c_col")
                        with col_ed2:
                            e_municipio = st.text_input("Municipio / Ciudad", value=cli_data.get("MUNICIPIO", ""), key="c_mun")
                            e_estado = st.text_input("Estado", value=cli_data.get("ESTADO", ""), key="c_est")
                            e_cp = st.text_input("Código Postal", value=cli_data.get("CP", ""), key="c_cp")

                        if st.button("Actualizar Cliente"):
                            datos_act_c = {"RFC": e_rfc, "RAZON_SOCIAL": e_razon_social, "FORMA_PAGO": e_forma_pago, "CALLE": e_calle, "NO_EXTERIOR": e_no_ext, "COLONIA": e_colonia, "MUNICIPIO": e_municipio, "ESTADO": e_estado, "CP": e_cp}
                            try:
                                if id_cli:
                                    supabase.table("clientes").update(datos_act_c).eq("id", id_cli).execute()
                                else:
                                    supabase.table("clientes").update(datos_act_c).eq("RAZON_SOCIAL", cli_data.get("RAZON_SOCIAL")).execute()
                                st.success("¡Cliente actualizado exitosamente!")
                                st.rerun()
                            except Exception as e: st.error(f"Error al actualizar: {e}")
                else: st.info("No hay clientes registrados.")
